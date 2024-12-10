########################################### 代码功能：进行模型验证 ##################################################
################################################## 导入模块 #########################################################

import pandas as pd
import matplotlib.pyplot as plt
from pandas import Series
import numpy as np
from sklearn import ensemble,preprocessing
from sklearn.preprocessing import minmax_scale,StandardScaler, Normalizer
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
from sklearn.feature_selection import SelectFromModel
import joblib
from sklearn.tree import DecisionTreeRegressor, export_graphviz
from sklearn import tree
from scipy.stats import pearsonr
import os
import sys
import time
from openpyxl import *

############################################### 输入要导入数据文件的名称 ###############################################
path_name       = os.getcwd()
file_name       = 'Randomly-split-data.py'
excel_name      = '2024-ML_20231228_all_scaler.xlsx'  # 导入的excel名称
columns_used_in_excel = [i for i in range(0,22)]                             # excel表中有数据的列
y_column_in_excel = 'lgk1'                                            # excel表中目标值那一列的title
cluster_name_column = 'clusters'                                        # excel表中团簇名称那一列的title
#######################################################################################################################


############################################设置划分数据集尝试次数####################################################
N = 50
count = 1
#######################################################################################################################
for i in range(0,N):
    print('数据集划分尝试次数：',i+1)
    os.chdir('{}'.format(path_name))
    ########################################### 定义新建文件夹函数 ########################################################
    def mkdir(path):
        # 去除首位空格
        path = path.strip()
        # 去除尾部 \ 符号
        path = path.rstrip("\\")
        # 判断路径是否存在
        # 存在     True
        # 不存在   False
        isExists = os.path.exists(path)
        # 判断结果
        if not isExists:
            # 如果不存在则创建目录
            # 创建目录操作函数
            os.makedirs(path)
    #        print(path + ' 创建成功')
            return True
        else:
            # 如果目录存在则不创建，并提示目录已存在
    #        print(path + ' 目录已存在')
            return False
    # 定义要创建的目录
    time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
    mkpath0 = "{}/not-good-splite-data".format(os.getcwd())
    mkpath = "{}/Splited-data-{}".format(path_name, time_str)
    mkdir(mkpath0)
    mkdir(mkpath)
    #######################################################################################################################

    ############################################ Define copy function ######################################################
    # srcfile 需要复制、移动的文件   
    # dstpath 目的地址
    import glob
    import shutil

    def mycopyfile(srcfile,dstpath):                       # 复制函数
        if not os.path.isfile(srcfile):
            print ("%s not exist!"%(srcfile))
        else:
            fpath,fname=os.path.split(srcfile)             # 分离文件名和路径
            if not os.path.exists(dstpath):
                os.makedirs(dstpath)                       # 创建路径
            shutil.copy(srcfile, dstpath + fname)          # 复制文件
            #print ("copy %s -> %s"%(srcfile, dstpath + fname))

    src_dir_info = '{}/'.format(path_name)

    dst_dir_info = '{}/Splited-data-{}/'.format(path_name, time_str)

    src_file_list =  glob.glob(src_dir_info + '{}'.format(excel_name))

    src_file_list2 =  glob.glob(src_dir_info + '{}'.format(file_name))

    for srcfile in src_file_list:
        mycopyfile(srcfile, dst_dir_info)

    for srcfile in src_file_list2:
        mycopyfile(srcfile, dst_dir_info)

    #######################################################################################################################

    print('Split data ...')

    ################################################## 导入数据 #########################################################
    os.chdir('{}'.format(path_name))
    df   = pd.read_excel('{}'.format(excel_name),usecols=columns_used_in_excel)    
    y    = df['{}'.format(y_column_in_excel)]
    X    = df.drop(['{}'.format(y_column_in_excel)], axis=1)

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1)  # 将数据划分为训练集和测试集
    
    save_path = '{}/Splited-data-{}'.format(path_name, time_str)

    X_train_output=pd.DataFrame(X_train)
    X_train_output.to_excel("{}/X_train_real.xlsx".format(save_path))
    y_train_output=pd.DataFrame(y_train)
    y_train_output.to_excel("{}/y_train_real.xlsx".format(save_path))

    X_test_output=pd.DataFrame(X_test)
    X_test_output.to_excel("{}/X_test_real.xlsx".format(save_path))
    y_test_output=pd.DataFrame(y_test)
    y_test_output.to_excel("{}/y_test_real.xlsx".format(save_path))
    
    
    mkpath_train = "{}/train_set".format(mkpath)
    mkpath_test = "{}/test_set".format(mkpath)
    mkdir(mkpath_train)
    mkdir(mkpath_test)

    shutil.move("{}/X_train_real.xlsx".format(save_path),mkpath_train)
    shutil.move("{}/y_train_real.xlsx".format(save_path),mkpath_train)

    shutil.move("{}/X_test_real.xlsx".format(save_path),mkpath_test)
    shutil.move("{}/y_test_real.xlsx".format(save_path),mkpath_test)

    os.chdir('{}'.format(mkpath_train))

    trian_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
    trian_data.to_excel('train_data.xlsx',encoding='gb18030')
    trian_data_path = "{}/train_data.xlsx".format(mkpath_train)
    wb = load_workbook(trian_data_path)
    ws = wb.active
    ws.delete_cols(1,2) #删除第 1 ,2列数据
    ws.delete_cols(len(columns_used_in_excel)) 
    wb.save(trian_data_path)
    mycopyfile(trian_data_path,"{}/".format(save_path))

    os.chdir('{}'.format(mkpath_test))
    test_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
    test_data.to_excel('test_data.xlsx',encoding='gb18030')
    test_data_path = "{}/test_data.xlsx".format(mkpath_test)
    wb = load_workbook(test_data_path)
    ws = wb.active
    ws.delete_cols(1,2) #删除第 1 ,2列数据
    ws.delete_cols(len(columns_used_in_excel)) 
    wb.save(test_data_path)
    mycopyfile(test_data_path,"{}/".format(save_path))



    os.chdir('{}'.format(path_name))
    src_file_list =  glob.glob("{}/".format(path_name) + '*.ipynb')
    for srcfile in src_file_list:
         mycopyfile(srcfile, "{}/".format(save_path))
    time.sleep(1)
    print('****************************The {}th good split data!***********************'.format(count))    
    count += 1

    #########################################判断测试集划分是否合理#############################################
#     # 设置分段
#     bins = [-15,-13,-12,-10,-7]

#     # 按分段离散化数据
#     segments=pd.cut(y_test,bins,right=False)

#     # 统计各分段数目
#     counts=pd.value_counts(segments,sort=False)

#     if counts.values[0] >2 and counts.values[1] >=4 and counts.values[2] >=1 and counts.values[3] >=1:
#         mkpath_train = "{}/train_set".format(mkpath)
#         mkpath_test = "{}/test_set".format(mkpath)
#         mkdir(mkpath_train)
#         mkdir(mkpath_test)
        
#         shutil.move("{}/X_train_real.xlsx".format(save_path),mkpath_train)
#         shutil.move("{}/y_train_real.xlsx".format(save_path),mkpath_train)
        
#         shutil.move("{}/X_test_real.xlsx".format(save_path),mkpath_test)
#         shutil.move("{}/y_test_real.xlsx".format(save_path),mkpath_test)
        
#         os.chdir('{}'.format(mkpath_train))

#         trian_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
#         trian_data.to_excel('train_data.xlsx',encoding='gb18030')
#         trian_data_path = "{}/train_data.xlsx".format(mkpath_train)
#         wb = load_workbook(trian_data_path)
#         ws = wb.active
#         ws.delete_cols(1,2) #删除第 1 ,2列数据
#         ws.delete_cols(len(columns_used_in_excel)) 
#         wb.save(trian_data_path)
#         mycopyfile(trian_data_path,"{}/".format(save_path))
        
#         os.chdir('{}'.format(mkpath_test))
#         test_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
#         test_data.to_excel('test_data.xlsx',encoding='gb18030')
#         test_data_path = "{}/test_data.xlsx".format(mkpath_test)
#         wb = load_workbook(test_data_path)
#         ws = wb.active
#         ws.delete_cols(1,2) #删除第 1 ,2列数据
#         ws.delete_cols(len(columns_used_in_excel)) 
#         wb.save(test_data_path)
#         mycopyfile(test_data_path,"{}/".format(save_path))
        
        
        
#         os.chdir('{}'.format(path_name))
#         src_file_list =  glob.glob("{}/".format(path_name) + '*.ipynb')
#         for srcfile in src_file_list:
#              mycopyfile(srcfile, "{}/".format(save_path))
#         time.sleep(1)
#         print('****************************The {}th good split data!***********************'.format(count))    
#         count += 1
        

#     else:
        
#         mkpath_train = "{}/train_set".format(mkpath)
#         mkpath_test = "{}/test_set".format(mkpath)
#         mkdir(mkpath_train)
#         mkdir(mkpath_test)
        
#         shutil.move("{}/X_train_real.xlsx".format(save_path),mkpath_train)
#         shutil.move("{}/y_train_real.xlsx".format(save_path),mkpath_train)
        
#         shutil.move("{}/X_test_real.xlsx".format(save_path),mkpath_test)
#         shutil.move("{}/y_test_real.xlsx".format(save_path),mkpath_test)
        
#         os.chdir('{}'.format(mkpath_train))

#         trian_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
#         trian_data.to_excel('train_data.xlsx',encoding='gb18030')
#         trian_data_path = "{}/train_data.xlsx".format(mkpath_train)
#         wb = load_workbook(trian_data_path)
#         ws = wb.active
#         ws.delete_cols(1,2) #删除第 1 ,2列数据
#         ws.delete_cols(len(columns_used_in_excel)) 
#         wb.save(trian_data_path)
#         mycopyfile("train_data.xlsx","{}/".format(save_path))
        
#         os.chdir('{}'.format(mkpath_test))
#         test_data = pd.concat(map(pd.read_excel, glob.glob('*.xlsx')),axis=1)
#         test_data.to_excel('test_data.xlsx',encoding='gb18030')
#         test_data_path = "{}/test_data.xlsx".format(mkpath_test)
#         wb = load_workbook(test_data_path)
#         ws = wb.active
#         ws.delete_cols(1,2) #删除第 1 ,2列数据
#         ws.delete_cols(len(columns_used_in_excel)) 
#         wb.save(test_data_path)
#         mycopyfile("test_data.xlsx","{}/".format(save_path))
        
#         os.chdir('{}'.format(path_name))
#         shutil.move(mkpath,mkpath0)
#         time.sleep(1)
#         print('Split data not good!')  

# ######################################################################################################################  

