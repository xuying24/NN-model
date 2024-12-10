#!/usr/bin/env python
# coding: utf-8

########################### 代码功能：利用GBR特征重要性和末尾淘汰机制进行特征选择####################################
################################################## 导入模块 #########################################################
import pandas as pd
import matplotlib.pyplot as plt
from pandas import Series
import numpy as np
from evolutionary_search import EvolutionaryAlgorithmSearchCV
from sklearn import ensemble, preprocessing,tree,svm
from sklearn.kernel_ridge import KernelRidge
from sklearn.tree import DecisionTreeRegressor, export_graphviz
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import minmax_scale, StandardScaler, Normalizer
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split, LeaveOneOut,GridSearchCV,KFold
from sklearn.metrics import r2_score
from scipy.stats import pearsonr
from sklearn.feature_selection import SelectFromModel
from collections import Counter
from pprint import pprint
import seaborn as sb
import os,random,joblib,linecache,sys,time
from openpyxl import load_workbook
os.system('bash -c "source /home/soft/anaconda3/etc/profile.d/conda.sh && conda activate base" ')
############################################**参数设置**################################################################
Feature_selection_LOOCV                  = 'ON'   # 如果选择非统计法，设置为ON，否则设置为OFF
Feature_selection_stat                   = 'ON'    # 如果选择统计法，设置为ON，否则设置为OFF
Model_selection                          = 'ON'    # 如果进行模型选择，设置为ON，否则设置为OFF
Model_selection_GBR                      = 'ON'    # 如果使用GBR（梯度提升回归）算法，设置为ON，否则设置为OFF
Model_selection_SVR                      = 'ON'    # 如果使用SVR（支持向量回归）算法，设置为ON，否则设置为OFF
Model_selection_BPANN                    = 'ON'    # 如果使用BPANN（反向传播神经网络）算法，设置为ON，否则设置为OFF
Model_selection_RF                                        = 'ON'    # 如果使用RF（随机森林）算法，设置为ON，否则设置为OFF
Model_selection_KRR                                     = 'ON'    # 如果使用KRR（核岭回归）算法，设置为ON，否则设置为OFF
Hyper_parameter_opt_and_Model_validation = 'ON'    # 如果进行超参数优化和模型验证，设置为ON，否则设置为OFF

Stat_feature_sele_cycles                 = 150     # 统计法进行特征选择时，设置循环次数
Model_selection_cycles                   = 1     # 设置模型选择循环次数，
stat_threshold_value                     = 15      # 统计法选择特征的阈值
#######################################################################################################################

############################################### 输入要导入数据文件的名称 ###############################################

path_name = os.getcwd()
print(path_name)
os.rename('train_data.xlsx', 'RhCo_train_data_for_feature_slection.xlsx')

file_name = 'RhCo--featrue_sele_r_method-loocv-stat.ipynb'
read_excel_name = 'RhCo_train_data_for_feature_slection.xlsx'  # 导入的excel名称
df_0 = pd.read_excel('{}'.format(read_excel_name))
df_0_head = df_0.columns.to_list()[0:]
columns_used_in_excel = list(range(len(df_0_head)))  # excel表中有数据的列
y_column_in_excel = df_0_head[-1]  # excel表中目标值那一列的title
cluster_name_column = df_0_head[0]  # excel表中团簇名称那一列的title
#######################################################################################################################

NUN_array = []
good_X_name_array = [NUN_array] * Stat_feature_sele_cycles  # 定义保存每次特征选择中最大 r 值对应的排好序的特征名的数组
good_name_weight_dict_array = [NUN_array] * Stat_feature_sele_cycles
start = time.time()

#########################################################################################################
def mkdir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)
    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)
        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False
    # 定义要创建的目录


#########################################################################################################

############################################ Define copy function ######################################################
# srcfile 需要复制、移动的文件
# dstpath 目的地址
import glob
import shutil


def mycopyfile(srcfile, dstpath):  # 复制函数
    if not os.path.isfile(srcfile):
        print("%s not exist!" % (srcfile))
    else:
        fpath, fname = os.path.split(srcfile)  # 分离文件名和路径
        if not os.path.exists(dstpath):
            os.makedirs(dstpath)  # 创建路径
        shutil.copy(srcfile, dstpath + fname)  # 复制文件
        print("copy %s -> %s" % (srcfile, dstpath + fname))


################################################创建文件夹和复制文件#################################################
time_str0 = time.strftime("%Y-%m-%d-%H-%M-%S")
mycopyfile('RhCo_train_data_for_feature_slection.xlsx', '{}/copy/'.format(path_name))
os.chdir('{}/copy'.format(path_name))
os.rename('RhCo_train_data_for_feature_slection.xlsx', 'train_data.xlsx')
mycopyfile('train_data.xlsx', '{}/'.format(path_name))
os.chdir('{}'.format(path_name))

#######################################################################################################################

# ******************************************************************************************************************************#


#######################################################非统计法特征选择#########################################################

if Feature_selection_LOOCV == 'ON':

    mkpath_fea_sele_loocv = "{}/feature_selection/loocv-{}-feature-selection".format(path_name, time.strftime("%Y-%m-%d-%H-%M-%S"))
    time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
    mkdir(mkpath_fea_sele_loocv)
    fp = open('{}/feature_selection/loocv-{}-feature-selection/feature-select-information.txt'.format(path_name, time_str),'a+', encoding='utf-8')

    ################################################## 导入数据 #########################################################
    df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel)
    print('Read excel name : {}\n'.format(read_excel_name), file=fp)
    X_name = df.columns.to_list()[1:-1]
    print('initial features in database:\n', file=fp)
    print(X_name, '\n', file=fp)
    y = df['{}'.format(y_column_in_excel)]
    X = df.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)
    initial_num_feature = len(X_name)

    ############################################## 定义存放数据的数组 #####################################################

    # ************************************* 存放【特征选择】过程中的重要数据 ***********************************************
    R2_scorel_feature = []  # 用于存放特征选择每次循环中的R2数值
    r_scorel_feature = []  # 用于存放特征选择每次循环中的r数值
    RMSE_scorel_feature = []  # 用于存放特征选择每次循环中的RMSE数值
    num_feature = []  # 用于存放特征选择每次循环中的当前特征数目
    r_scorel_arry = [NUN_array] * (initial_num_feature)
    features_arry = [NUN_array] * (initial_num_feature)
    # **********************************************************************************************************************

    ################################################### 特征选择 ##########################################################
    print('#################################################### Feature selection ##################################################',file=fp)
    # ***************************************** 利用循环语句删除不重要特征 #################################################
    Min_num_featrue = 1
    num_cycle = (initial_num_feature - Min_num_featrue) + 1  # 设置循环次数
    for i in range(0, num_cycle):
        X_name_sorted_idx = []

        X_dict = X.values
        y_dict = y.values

        print('####################################################{} features to selected ##################################################'.format(initial_num_feature - i), file=fp)
        X_name_sorted_idx = []  # 定义存放根据特征重要性排好序的特征名的数组
        y_test_real_all = []
        y_test_predict_all = []

        loo = LeaveOneOut()
        loo.get_n_splits(X_dict)
        for train_index, test_index in loo.split(X_dict):
            X_train_loo, X_test_loo = X_dict[train_index], X_dict[test_index]
            y_train_loo, y_test_loo = y_dict[train_index], y_dict[test_index]
            y_test_real_all.append(y_test_loo)

            reg = ensemble.GradientBoostingRegressor()  # 定义模型名称
            reg.fit(X_train_loo, y_train_loo)  # 模型拟合
            y_test_predict = reg.predict(X_test_loo)
            y_test_predict_all.append(y_test_predict)

        y_test_real_all = np.squeeze(y_test_real_all)
        y_test_predict_all = np.squeeze(y_test_predict_all)

        R2 = r2_score(y_test_real_all, y_test_predict_all)  # 在测试集上计算决定系数
        r, _ = pearsonr(y_test_real_all, y_test_predict_all)  # 在测试集上计算皮尔逊相关系数
        r_scorel_arry[i] = r

        RMSE = np.sqrt(mean_squared_error(y_test_real_all, y_test_predict_all))  # 在测试集上计算均方根误差

        print("The coefficient of R2 (including {} features) : {:.4f}".format((initial_num_feature - i), R2), file=fp)
        print("The coefficient of pearsonr (including {} features): {:.4f}".format((initial_num_feature - i), r),file=fp)
        print("The RMSE(including {} features): {:.4f}".format((initial_num_feature - i), RMSE), file=fp)

        feature_importance = reg.feature_importances_  # 利用GBR方法计算得到各特征的重要性
        sorted_idx = np.argsort(feature_importance)  # 利用GBR方法计算得到各特征重要性的排序索引
        feature_importance_sorted = feature_importance[sorted_idx]

        for j in sorted_idx:
            X_name_sorted_idx.append(X_name[j])
        features_arry[i] = X_name_sorted_idx

        num_feature.append(X.shape[1])
        R2_scorel_feature.append(R2)
        r_scorel_feature.append(r)
        ############################################## 绘制特征重要性的条形图 #########################################
                
        pos = np.arange(sorted_idx.shape[0]) + 0.5
        plt.clf()
        fig = plt.figure(figsize=(12, 12))
        plt.subplot(1, 1, 1)
        plt.barh(pos, feature_importance_sorted, height=0.8, align="center")
        plt.xticks(fontproperties='Times New Roman', size=14)
        plt.yticks(pos, np.array(X_name)[sorted_idx], fontproperties='Times New Roman', size=14)
        plt.title("Feature Importance (MDI): r = {:.4f}, R2 = {:.4f}, RMSE = {:.4f}".format(r, R2, RMSE),fontproperties='Times New Roman', size=16)
        plt.savefig('{}/feature_selection/loocv-{}-feature-selection/Feature_Importance_{}_features.png'.format(path_name,time_str, (initial_num_feature - i)))

        ######################################### 绘制所有特征之间相关性的热力图 #########################################
        plt.clf()
        plt.subplot(1, 1, 1)
        plt.xticks(fontproperties='Times New Roman', size=14)
        plt.yticks(fontproperties='Times New Roman', size=14)
        # dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlGnBu", annot=True)
        dataplot = sb.heatmap(X.corr(method='pearson'), cmap="YlGnBu", annot=True)  # 蓝绿色
        # dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlOrRd")#红橙色
        plt.savefig('{}/feature_selection/loocv-{}-feature-selection/{}_features_heatmap.png'.format(path_name, time_str,(initial_num_feature - i)))
        #######################################################################################################################

        print('Selected_features:', X_name, file=fp)
        X.drop(columns=X.columns[[sorted_idx[0]]], inplace=True)  # 特征按照重要性进行排序后，将重要性最低的特征从数据列表中删除
        del X_name[sorted_idx[0]]  # 将对应的特征名称一并删除
        # **********************************************************************************************************************

    #######################################################################################################################

    ################################### 绘制决定系数和皮尔逊相关系数与特征数目的关系图 ####################################
    fig = plt.figure(figsize=(12, 4))
    plt.subplot(1, 1, 1)
    plt.xticks(fontproperties='Times New Roman', size=14)
    plt.yticks(fontproperties='Times New Roman', size=14)
    plt.xlabel('Feature number', fontproperties='Times New Roman', size=14)
    plt.ylabel('R2 of GBR algorithm', fontproperties='Times New Roman', size=14)
    plt.plot(num_feature, R2_scorel_feature)
    plt.savefig('{}/feature_selection/loocv-{}-feature-selection/R2_vs_Feature_number_{}.png'.format(path_name, time_str,time_str))
    plt.clf()
    plt.subplot(1, 1, 1)
    plt.xticks(fontproperties='Times New Roman', size=14)
    plt.yticks(fontproperties='Times New Roman', size=14)
    plt.xlabel('Feature number', fontproperties='Times New Roman', size=14)
    plt.ylabel('r of GBR algorithm', fontproperties='Times New Roman', size=14)
    plt.plot(num_feature, r_scorel_feature)
    plt.savefig('{}/feature_selection/loocv-{}-feature-selection/r_vs_Feature_number_{}.png'.format(path_name, time_str,time_str))
    fp.close()
    plt.cla()
    plt.close("all")
    ########################################################################################################################
    print('*** Feature selection finished ***')
    fp.close()

    ###########################################生成【RhCo_for_model_selection.xisx】#########################################
    # 读取非统计法选择的特征
    selected_featu = features_arry[r_scorel_arry.index(max(r_scorel_arry))]
    sele_head = ['{}'.format(cluster_name_column)] + selected_featu + ['{}'.format(y_column_in_excel)]

    # 选择Excel表中的重要特征列，然后生成RhCo_for_model_selection.xisx
    os.chdir('{}'.format(path_name))
    df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel, sheet_name='Sheet1')
    all_head_name = df.columns.to_list()[0:]
    df_1 = pd.read_excel('{}'.format(read_excel_name), usecols=sele_head)
    sele_fea_output = pd.DataFrame(df_1)
    sele_fea_output.to_excel("{}/RhCo_for_model_selection.xlsx".format(path_name))
    model_selection_excel = "{}/RhCo_for_model_selection.xlsx".format(path_name)
    wb = load_workbook(model_selection_excel)
    ws = wb.active
    ws.delete_cols(1)  # 删除第 1列数据
    wb.save(model_selection_excel)
 
    #############################################################################################################################

#######################################################统计法特征选择#########################################################

if Feature_selection_stat == 'ON':

    for k in range(0, Stat_feature_sele_cycles):
        ################################################## 导入数据 #########################################################

        df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel)

        X_name = df.columns.to_list()[1:-1]

        y = df['{}'.format(y_column_in_excel)]
        X = df.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)

        initial_num_feature = len(X_name)

        ######################################################################################################################

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1)  # 将数据划分为训练集和测试集

        ########################################### 定义新建文件夹函数 ########################################################
        mkpath_featu_sele_stat = "{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection".format(path_name, time_str0,time.strftime("%Y-%m-%d-%H-%M-%S"))
        time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
        mkdir(mkpath_featu_sele_stat)
        #######################################################################################################################

        fp = open('{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection/feature-select-information.txt'.format(path_name, time_str0, time_str), 'a+')

        #######################################################################################################################

        ############################################## 定义存放数据的数组 #####################################################

        # ************************************* 存放【特征选择】过程中的重要数据 ***********************************************
        R2_scorel_feature = []  # 用于存放特征选择每次循环中的R2数值
        r_scorel_feature = []  # 用于存放特征选择每次循环中的r数值
        RMSE_scorel_feature = []  # 用于存放特征选择每次循环中的RMSE数值
        num_feature = []  # 用于存放特征选择每次循环中的当前特征数目
        X_name_array = [NUN_array] * 1000
        Name_weight_dict_array = [NUN_array] * 1000

        # **********************************************************************************************************************

        ################################################### 特征选择 ##########################################################
        print('#################################################### Feature selection ##################################################',file=fp)
        # ***************************************** 利用循环语句删除不重要特征 #################################################
        Min_num_featrue = 1
        num_cycle = (initial_num_feature - Min_num_featrue) + 1  # 设置循环次数

        for i in range(0, num_cycle):
            X_dict = X_train.values
            y_dict = y_train.values
            print('####################################################{} features to selected ##################################################'.format(initial_num_feature - i), file=fp)
            X_name_sorted_idx = []  # 定义存放根据特征重要性排好序的特征名的数组
            y_test_real_all = []
            y_test_predict_all = []

            loo = LeaveOneOut()
            loo.get_n_splits(X_dict)
            for train_index, test_index in loo.split(X_dict):
                X_train_loo, X_test_loo = X_dict[train_index], X_dict[test_index]
                y_train_loo, y_test_loo = y_dict[train_index], y_dict[test_index]
                y_test_real_all.append(y_test_loo)

                reg = ensemble.GradientBoostingRegressor(random_state=90)  # 定义模型名称
                reg.fit(X_train_loo, y_train_loo)  # 模型拟合
                y_test_predict = reg.predict(X_test_loo)
                y_test_predict_all.append(y_test_predict)

            y_test_real_all = np.squeeze(y_test_real_all)
            y_test_predict_all = np.squeeze(y_test_predict_all)

            R2 = r2_score(y_test_real_all, y_test_predict_all)  # 在测试集上计算决定系数
            r, _ = pearsonr(y_test_real_all, y_test_predict_all)  # 在测试集上计算皮尔逊相关系数

            RMSE = np.sqrt(mean_squared_error(y_test_real_all, y_test_predict_all))  # 在测试集上计算均方根误差

            print("The coefficient of R2 (including {} features) : {:.4f}".format((initial_num_feature - i), R2),file=fp)
            print("The coefficient of pearsonr (including {} features): {:.4f}".format((initial_num_feature - i), r),file=fp)
            print("The RMSE(including {} features): {:.4f}".format((initial_num_feature - i), RMSE), file=fp)

            feature_importance = reg.feature_importances_  # 利用GBR方法计算得到各特征的重要性
            sorted_idx = np.argsort(feature_importance)  # 利用GBR方法计算得到各特征重要性的排序索引
            feature_importance_sorted = feature_importance[sorted_idx]

            num_feature.append(X_train.shape[1])
            R2_scorel_feature.append(R2)
            r_scorel_feature.append(r)

            for jj in sorted_idx:
                X_name_sorted_idx.append(X_name[jj])  # 将特征名根据重要性进行排序
            X_name_array[(len(r_scorel_feature) - 1)] = X_name_sorted_idx  # 存放每次循环中排好序的特征名

            keys = X_name_sorted_idx
            values = feature_importance_sorted.tolist()
            Name_weight_dict_array[(len(r_scorel_feature) - 1)] = dict(zip(keys, values))

            ############################################## 绘制特征重要性的条形图 #########################################

            pos = np.arange(sorted_idx.shape[0]) + 0.5
            plt.clf()
            fig = plt.figure(figsize=(12, 12))
            plt.subplot(1, 1, 1)
            plt.barh(pos, feature_importance_sorted, height=0.8, align="center")
            plt.xticks(fontproperties='Times New Roman', size=14)
            plt.yticks(pos, np.array(X_name)[sorted_idx], fontproperties='Times New Roman', size=14)
            plt.title("Feature Importance (MDI): r = {:.4f}, R2 = {:.4f}, RMSE = {:.4f}".format(r, R2, RMSE),fontproperties='Times New Roman', size=16)
            plt.savefig('{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection/Feature_Importance_{}_features.png'.format(path_name, time_str0, time_str, (initial_num_feature - i)))

            ######################################### 绘制所有特征之间相关性的热力图 #########################################

            plt.clf()
            plt.subplot(1, 1, 1)
            plt.xticks(fontproperties='Times New Roman', size=14)
            plt.yticks(fontproperties='Times New Roman', size=14)
            # dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlGnBu", annot=True)
            dataplot = sb.heatmap(X_train.corr(method='pearson'), cmap="YlGnBu")  # 蓝绿色
            # dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlOrRd")#红橙色
            plt.savefig('{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection/{}_features_heatmap.png'.format(path_name, time_str0, time_str, (initial_num_feature - i)))
          
            #######################################################################################################################

            #X_train.drop(columns=X_train.columns[[sorted_idx[0]]],inplace = True)                        # 特征按照重要性进行排序后，将重要性最低的特征从数据列表中删除
            X_train = X_train.drop(columns=X_train.columns[[sorted_idx[0]]])  # 特征按照重要性进行排序后，将重要性最低的特征从数据列表中删除
            del X_name[sorted_idx[0]]  # 将对应的特征名称一并删除

        # **********************************************************************************************************************

        #######################################################################################################################

        ################################### 绘制决定系数和皮尔逊相关系数与特征数目的关系图 ####################################
        fig = plt.figure(figsize=(12, 4))
        plt.subplot(1, 1, 1)
        plt.xticks(fontproperties='Times New Roman', size=14)
        plt.yticks(fontproperties='Times New Roman', size=14)
        plt.xlabel('Feature number', fontproperties='Times New Roman', size=14)
        plt.ylabel('R2 of GBR algorithm', fontproperties='Times New Roman', size=14)
        plt.plot(num_feature, R2_scorel_feature)
        plt.savefig('{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection/R2_vs_Feature_number.png'.format(path_name, time_str0, time_str))
        plt.clf()
       
        plt.figure(figsize=[12, 4])
        plt.subplot(1, 1, 1)
        plt.xticks(fontproperties='Times New Roman', size=14)
        plt.yticks(fontproperties='Times New Roman', size=14)
        plt.xlabel('Feature number', fontproperties='Times New Roman', size=14)
        plt.ylabel('r of GBR algorithm', fontproperties='Times New Roman', size=14)
        plt.plot(num_feature, r_scorel_feature)
        plt.savefig('{}/feature_selection/{}-feature-selection/loocv-{}-feature-selection/r_vs_Feature_number.png'.format(path_name, time_str0, time_str))

        good_X_name_array[k] = (
        X_name_array[r_scorel_feature.index(max(r_scorel_feature))])  # 保存本次特征选择中最大 r 值对应的排好序的特征名
        good_name_weight_dict_array[k] = (Name_weight_dict_array[r_scorel_feature.index(max(r_scorel_feature))])
        fp.close()
        plt.cla()
        plt.close("all")

    fp1 = open('{}/feature_selection/{}-feature-selection/{}-cycles-feature-select-information.txt'.format(path_name,time_str0,Stat_feature_sele_cycles), 'a+',encoding='utf-8')

    df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel)
    X_name = df.columns.to_list()[1:-1]
    print('All features:{}\n'.format(X_name), file=fp1)
    print('\n', file=fp1)
    print('\n', file=fp1)

    print('good_X_name_array【{}次特征选择筛选得到的二维特征列表】:'.format(Stat_feature_sele_cycles), file=fp1)
    print(good_X_name_array, file=fp1)
    print('\n', file=fp1)

    '''100次筛选得到的特征存放在二维列表（good_X_name_array）中，将二维列表转化为一维列表（good_X_name_array_1_dim）'''
    good_X_name_array_1_dim = [n for item in good_X_name_array for n in item]
    print('good_X_name_array_1_dim【{}次特征选择筛选得到的一维特征列表】:'.format(Stat_feature_sele_cycles), file=fp1)
    print(good_X_name_array_1_dim, file=fp1)
    print('\n', file=fp1)
    ########################################################################################################################################

    '''统计100次特征筛选中，各个特征出现的次数'''
    list = dict(Counter(good_X_name_array_1_dim))
    stat_features = {key: value for key, value in list.items() if value > 0}
    stat_features_order = dict(sorted(stat_features.items(), key=lambda x: x[1], reverse=True))
    print('stat_features_order【统计各特征出现次数，并从大到小排序】:', file=fp1)
    for key, value in stat_features_order.items():
        print(key, ':', value, file=fp1)
    print('\n', file=fp1)
    ########################################################################################################################################

    '''100次筛选得到的特征重要性权重存放在二维字典（good_name_weight_dict_array）中，
    对各个特征在100次筛选中的重要性权重进行求和放入字典（sum_dict）中，并按从大到小排好序和求各特征重要性权重的平均值'''
    sum_dict = dict()
    for n in good_name_weight_dict_array:
        for key, value in n.items():
            if key in sum_dict:
                sum_dict[key] += value
            else:
                sum_dict[key] = value

    sum_dict_order = dict(sorted(sum_dict.items(), key=lambda x: x[1], reverse=True))
    print('sum_dict_order【各特征的重要性权重值求和，并从大到小排序】:', file=fp1)
    for key, value in sum_dict_order.items():
        print(key, ':', value, file=fp1)
    print('\n', file=fp1)

    sum_dict_order_avg = dict()
    for key, value in sum_dict_order.items():
        sum_dict_order_avg[key] = value / Stat_feature_sele_cycles
    print('sum_dict_order_avg【各特征的平均重要性权重值，并从大到小排序】:', file=fp1)
    for key, value in sum_dict_order_avg.items():
        print(key, ':', value, file=fp1)
    print('\n', file=fp1)
    ########################################################################################################################################

    '''将统计得到的特征次数*该特征对应的平均权重'''
    featu_multi_weight = dict()
    for key, value in stat_features_order.items():
        if key in sum_dict_order_avg:
            featu_multi_weight[key] = sum_dict_order_avg[key] * value

    featu_multi_weight_order = dict(sorted(featu_multi_weight.items(), key=lambda x: x[1], reverse=True))
    print('featu_multi_weight_order【特征次数*该特征对应的平均权重，并从大到小排序】:', file=fp1)
    for key, value in featu_multi_weight_order.items():
        print(key, ':', value, file=fp1)
    print('\n', file=fp1)
    ############################################################################################################################################
    end = time.time()

    print('run time:{} s'.format(str(end - start)), file=fp1)
    print('run time:{} min'.format(str((end - start) / 60)), file=fp1)

    print('*** Feature selection finished ***')

    fp1.close()
    del good_X_name_array, good_name_weight_dict_array

    ###########################################生成【RhCo_for_model_selection.xisx】#########################################
    # 读取统计法选择的特征
    os.chdir('{}/feature_selection/{}-feature-selection'.format(path_name, time_str0))
    txt_path = '{}-cycles-feature-select-information.txt'.format(Stat_feature_sele_cycles)
    file = open(txt_path, "r", encoding='utf-8')
    txt = file.readlines()
    file.close()

    os.chdir('{}'.format(path_name))
    df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel)
    os.chdir('{}/feature_selection/{}-feature-selection'.format(path_name, time_str0))
    X_name = df.columns.to_list()[1:-1]
    initial_num_feature = len(X_name)

    line_num = 0
    selected_featu = []
    all_featu_names = []
    for line in txt:
        line_num += 1
        if line.find('featu_multi_weight_order') > -1:
            important_line_head = line_num + 1
            important_line_tail = line_num + initial_num_feature - 3
            for line_line in range(important_line_head, important_line_tail):
                print(float(linecache.getline(txt_path, line_line).split(':')[1]))
                if float(linecache.getline(txt_path, line_line).split(':')[1]) >= stat_threshold_value:
                    selected_featu = selected_featu + [linecache.getline(txt_path, line_line).split()[0]]
    sele_head = ['{}'.format(cluster_name_column)] + selected_featu + ['{}'.format(y_column_in_excel)]

    # 选择Excel表中的重要特征列，然后生成RhCo_for_model_selection.xisx
    os.chdir('{}'.format(path_name))
    df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel, sheet_name='Sheet1')
    all_head_name = df.columns.to_list()[0:]
    df_1 = pd.read_excel('{}'.format(read_excel_name), usecols=sele_head)
    sele_fea_output = pd.DataFrame(df_1)
    sele_fea_output.to_excel("{}/RhCo_for_model_selection.xlsx".format(path_name))
    model_selection_excel = "{}/RhCo_for_model_selection.xlsx".format(path_name)
    wb = load_workbook(model_selection_excel)
    ws = wb.active
    ws.delete_cols(1)  # 删除第 1列数据
    wb.save(model_selection_excel)

    ###########################################################################################################################


if Model_selection == 'ON':
    start = time.time()

    ################################################### 定义数组 ##########################################################
    R_GBR                                             = []
    RMSE_GBR                                          = []
    R_GBR_test_set                                    = []
    RMSE_GBR_test_set                                 = []
  
    R_SVR                                             = []
    RMSE_SVR                                          = []
    R_SVR_test_set                                    = []
    RMSE_SVR_test_set                                 = []
  
    R_BPANN                                           = []
    RMSE_BPANN                                        = []
    R_BPANN_test_set                                  = []
    RMSE_BPANN_test_set                               = []
  
    R_RF                                              = []
    RMSE_RF                                           = []
    R_RF_test_set                                     = []
    RMSE_RF_test_set                                  = []
  
    R_KRR                                             = []
    RMSE_KRR                                          = []
    R_KRR_test_set                                    = []
    RMSE_KRR_test_set                                 = [] 
    #######################################################################################################################

    ############################################### 输入要导入数据文件的名称 ###############################################

    read_excel_name_1 = 'RhCo_for_model_selection.xlsx'  # 导入的excel名称
    df_1 = pd.read_excel('{}'.format(read_excel_name_1))
    df_1_head = df_1.columns.to_list()[0:]
    #print(df_1_head)
    columns_used_in_excel = list(range(len(df_1_head)))  # excel表中有数据的列
    #######################################################################################################################

    mkpath_model_sele = "{}/Model_selection/{}-features-{}-model-selection-{}-cycle".format(path_name, columns_used_in_excel[-2],time.strftime("%Y-%m-%d-%H-%M-%S"), Model_selection_cycles)
    time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
    mkdir(mkpath_model_sele)
    #######################################################################################################################

    fp = open('{}/Model_selection/{}-features-{}-model-selection-{}-cycle/model-selection-information.txt'.format(path_name,columns_used_in_excel[-2],time_str,Model_selection_cycles), 'a+')

    #######################################################################################################################

    print('Model selection ...')

    ################################################## 导入数据 #########################################################

    df = pd.read_excel('{}'.format(read_excel_name_1), usecols=columns_used_in_excel)
    y = df['{}'.format(y_column_in_excel)]
    X = df.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)
    X = X.values
    y = y.values
    ############################################### n-estimators ########################################################

    for i in range(0, Model_selection_cycles):
        print('Cycle {} of model selection'.format(i + 1))
        y_test_real_loo      = []
        y_test_predict_GBR   = []
        y_test_predict_SVR   = []
        y_test_predict_BPANN = []
        y_test_predict_RF    = []
        y_test_predict_KRR   = []

        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.1)  # 将数据划分为训练集和测试集

        loo = LeaveOneOut()
        loo.get_n_splits(X_train)
        for train_index, test_index in loo.split(X_train):
            X_train_loo, X_test_loo = X_train[train_index], X_train[test_index]
            y_train_loo, y_test_loo = y_train[train_index], y_train[test_index]

            y_test_real_loo.append(y_test_loo)
            
            if Model_selection_GBR == 'ON': 
               reg = ensemble.GradientBoostingRegressor(random_state=90)
               GBR = reg.fit(X_train_loo, y_train_loo)
               y_test_predict = GBR.predict(X_test_loo)
               y_test_predict_GBR.append(y_test_predict)

            if Model_selection_SVR == 'ON':
               SVR_reg = svm.SVR()
               SVR = SVR_reg.fit(X_train_loo, y_train_loo)
               y_test_predict = SVR.predict(X_test_loo)
               y_test_predict_SVR.append(y_test_predict)

            if Model_selection_BPANN == 'ON':
               BPANN = MLPRegressor(random_state=90, max_iter=4000).fit(X_train_loo, y_train_loo)
               y_test_predict = BPANN.predict(X_test_loo)
               y_test_predict_BPANN.append(y_test_predict)
 
            if Model_selection_RF == 'ON':
               RF = ensemble.RandomForestRegressor(random_state=90).fit(X_train_loo, y_train_loo)
               y_test_predict = RF.predict(X_test_loo)
               y_test_predict_RF.append(y_test_predict)

            if Model_selection_KRR == 'ON':      
               KRR = KernelRidge(alpha=1.0).fit(X_train_loo, y_train_loo)
               y_test_predict = KRR.predict(X_test_loo)
               y_test_predict_KRR.append(y_test_predict)

        y_test_real_loo = np.squeeze(y_test_real_loo)

        if Model_selection_GBR == 'ON':
           print('********************************************* GBR of cycle {} ****************************************************'.format(i + 1), file=fp)
           y_test_predict_GBR = np.squeeze(y_test_predict_GBR)
           r_GBR,_ = pearsonr(y_test_real_loo, y_test_predict_GBR)
           R_GBR.append(r_GBR)
           rmse_GBR = np.sqrt(mean_squared_error(y_test_real_loo,y_test_predict_GBR))
           RMSE_GBR.append(rmse_GBR)
           print('r_GBR:{}'.format(r_GBR), file=fp)
           print('R_GBR(All cycles from now):{}'.format(R_GBR), file=fp)
           print('RMSE_GBR:{}'.format(RMSE_GBR),file=fp)
           print('RMSE_GBR(All cycles from now):{}'.format(RMSE_GBR),file=fp)

           reg = ensemble.GradientBoostingRegressor(random_state=90)
           GBR = reg.fit(X_train, y_train)
           r_GBR_test,_ = pearsonr(y_test, GBR.predict(X_test))
           R_GBR_test_set.append(r_GBR_test)
           RMSE_GBR_test = np.sqrt(mean_squared_error(y_test,GBR.predict(X_test)))
           RMSE_GBR_test_set.append(RMSE_GBR_test)
           print('r_GBR_test:{}'.format(r_GBR_test), file=fp)
           print('R_GBR_test_set(All cycles from now):{}'.format(R_GBR_test_set), file=fp)
           print('RMSE_GBR_test:{}'.format(RMSE_GBR_test),file=fp)
           print('RMSE_GBR_test_set(All cycles from now):{}'.format(RMSE_GBR_test_set),file=fp)

        if Model_selection_SVR == 'ON':
           print('********************************************* SVR of cycle {} ****************************************************'.format(i+1),file=fp)
           y_test_predict_SVR = np.squeeze(y_test_predict_SVR)
           r_SVR,_ = pearsonr(y_test_real_loo,y_test_predict_SVR)
           R_SVR.append(r_SVR)
           rmse_SVR = np.sqrt(mean_squared_error(y_test_real_loo,y_test_predict_SVR))
           RMSE_SVR.append(rmse_SVR)
           print('r_SVR:{}'.format(r_SVR),file=fp)
           print('R_SVR(All cycles from now):{}'.format(R_SVR),file=fp)
           print('RMSE_SVR:{}'.format(RMSE_SVR),file=fp)
           print('RMSE_SVR(All cycles from now):{}'.format(RMSE_SVR),file=fp)

           SVR_reg = svm.SVR()
           SVR = SVR_reg.fit(X_train, y_train)
           r_SVR_test,_ = pearsonr(y_test,SVR.predict(X_test))
           R_SVR_test_set.append(r_SVR_test)
           RMSE_SVR_test = np.sqrt(mean_squared_error(y_test,SVR.predict(X_test)))
           RMSE_SVR_test_set.append(RMSE_SVR_test)
           print('r_SVR_test:{}'.format(r_SVR_test),file=fp)
           print('R_SVR_test_set(All cycles from now):{}'.format(R_SVR_test_set),file=fp)
           print('RMSE_SVR_test:{}'.format(RMSE_SVR_test),file=fp)
           print('RMSE_SVR_test_set(All cycles from now):{}'.format(RMSE_SVR_test_set),file=fp)

        if Model_selection_BPANN == 'ON':
           print('********************************************* BPANN of cycle {} ****************************************************'.format(i+1),file=fp)
           y_test_predict_BPANN = np.squeeze(y_test_predict_BPANN)
           r_BPANN,_ = pearsonr(y_test_real_loo,y_test_predict_BPANN)
           R_BPANN.append(r_BPANN)
           rmse_BPANN = np.sqrt(mean_squared_error(y_test_real_loo,y_test_predict_BPANN))
           RMSE_BPANN.append(rmse_BPANN)
           print('r_BPANN:{}'.format(r_BPANN),file=fp)
           print('R_BPANN(All cycles from now):{}'.format(R_BPANN),file=fp)
           print('RMSE_BPANN:{}'.format(RMSE_BPANN),file=fp)
           print('RMSE_BPANN(All cycles from now):{}'.format(RMSE_BPANN),file=fp)


           BPANN = MLPRegressor(random_state=90, max_iter=4000).fit(X_train, y_train)
           r_BPANN_test,_ = pearsonr(y_test,BPANN.predict(X_test))
           R_BPANN_test_set.append(r_BPANN_test)
           RMSE_BPANN_test = np.sqrt(mean_squared_error(y_test,BPANN.predict(X_test)))
           RMSE_BPANN_test_set.append(RMSE_BPANN_test)
           print('r_BPANN_test:{}'.format(r_BPANN_test),file=fp)
           print('R_BPANN_test_set(All cycles from now):{}'.format(R_BPANN_test_set),file=fp)
           print('RMSE_BPANN_test:{}'.format(RMSE_BPANN_test),file=fp)
           print('RMSE_BPANN_test_set(All cycles from now):{}'.format(RMSE_BPANN_test_set),file=fp)

        if Model_selection_RF == 'ON':
           print('********************************************* RF of cycle {} ****************************************************'.format(i+1),file=fp)
           y_test_predict_RF = np.squeeze(y_test_predict_RF)
           r_RF,_ = pearsonr(y_test_real_loo,y_test_predict_RF)
           R_RF.append(r_RF)
           rmse_RF = np.sqrt(mean_squared_error(y_test_real_loo,y_test_predict_RF))
           RMSE_RF.append(rmse_RF)
           print('r_RF:{}'.format(r_RF),file=fp)
           print('R_RF(All cycles from now):{}'.format(R_RF),file=fp)
           print('RMSE_RF:{}'.format(RMSE_RF),file=fp)
           print('RMSE_RF(All cycles from now):{}'.format(RMSE_RF),file=fp)

           RF = ensemble.RandomForestRegressor(random_state=90).fit(X_train, y_train)
           r_RF_test,_ = pearsonr(y_test,RF.predict(X_test))
           R_RF_test_set.append(r_RF_test)
           RMSE_RF_test = np.sqrt(mean_squared_error(y_test,RF.predict(X_test)))
           RMSE_RF_test_set.append(RMSE_RF_test)
           print('r_RF_test:{}'.format(r_RF_test),file=fp)
           print('R_RF_test_set(All cycles from now):{}'.format(R_RF_test_set),file=fp)
           print('RMSE_RF_test:{}'.format(RMSE_RF_test),file=fp)
           print('RMSE_RF_test_set(All cycles from now):{}'.format(RMSE_RF_test_set),file=fp)


        if Model_selection_RF == 'ON':
           print('************************** KRR of cycle {} ****************************************************'.format(i+1),file=fp)
           y_test_predict_KRR = np.squeeze(y_test_predict_KRR)
           r_KRR,_ = pearsonr(y_test_real_loo,y_test_predict_KRR)
           rmse_KRR = np.sqrt(mean_squared_error(y_test_real_loo,y_test_predict_KRR))
           R_KRR.append(r_KRR)
           RMSE_KRR.append(rmse_KRR)
           print('r_KRR:{}'.format(r_KRR),file=fp)
           print('R_KRR(All cycles from now):{}'.format(R_KRR),file=fp)
           print('RMSE_KRR:{}'.format(RMSE_KRR),file=fp)
           print('RMSE_KRR(All cycles from now):{}'.format(RMSE_KRR),file=fp)

           KRR = KernelRidge(alpha=1.0).fit(X_train, y_train)
           r_KRR_test,_ = pearsonr(y_test,KRR.predict(X_test))
           R_KRR_test_set.append(r_KRR_test)
           RMSE_KRR_test = np.sqrt(mean_squared_error(y_test,KRR.predict(X_test)))
           RMSE_KRR_test_set.append(RMSE_KRR_test)
           print('r_KRR_test:{}'.format(r_KRR_test),file=fp)
           print('R_KRR_test_set(All cycles from now):{}'.format(R_KRR_test_set),file=fp)
           print('RMSE_KRR_test:{}'.format(RMSE_KRR_test),file=fp)
           print('RMSE_KRR_test_set(All cycles from now):{}'.format(RMSE_KRR_test_set),file=fp)


    if Model_selection_GBR == 'ON':
       R_GBR_avg_100_times = np.mean(R_GBR)
       RMSE_GBR_avg_100_times = np.mean(RMSE_GBR)
       R_GBR_test_set = np.mean(R_GBR_test_set)
       RMSE_GBR_test_set = np.mean(RMSE_GBR_test_set)

    if Model_selection_SVR == 'ON':
       R_SVR_avg_100_times = np.mean(R_SVR)
       RMSE_SVR_avg_100_times = np.mean(RMSE_SVR)
       R_SVR_test_set = np.mean(R_SVR_test_set)
       RMSE_SVR_test_set = np.mean(RMSE_SVR_test_set)

    if Model_selection_BPANN == 'ON':
       R_BPANN_avg_100_times = np.mean(R_BPANN)
       RMSE_BPANN_avg_100_times = np.mean(RMSE_BPANN)
       R_BPANN_test_set = np.mean(R_BPANN_test_set)
       RMSE_BPANN_test_set = np.mean(RMSE_BPANN_test_set)

    if Model_selection_RF == 'ON':
       R_RF_avg_100_times = np.mean(R_RF)
       RMSE_RF_avg_100_times = np.mean(RMSE_RF)
       R_RF_test_set = np.mean(R_RF_test_set)
       RMSE_RF_test_set = np.mean(RMSE_RF_test_set)
  
    if Model_selection_KRR == 'ON':
       R_KRR_avg_100_times = np.mean(R_KRR)
       RMSE_KRR_avg_100_times = np.mean(RMSE_KRR)
       R_KRR_test_set = np.mean(R_KRR_test_set)
       RMSE_KRR_test_set = np.mean(RMSE_KRR_test_set)

    print('********************************************* summry ****************************************************',file=fp)
    print('Trianing set:\n', file=fp)
    if Model_selection_GBR == 'ON':
       print('R_GBR_avg_100_times:{}'.format(R_GBR_avg_100_times), file=fp)
       print('RMSE_GBR_avg_100_times:{}\n'.format(RMSE_GBR_avg_100_times),file=fp)
    if Model_selection_SVR == 'ON':
       print('R_SVR_avg_100_times:{}'.format(R_SVR_avg_100_times),file=fp)
       print('RMSE_SVR_avg_100_times:{}\n'.format(RMSE_SVR_avg_100_times),file=fp)
    if Model_selection_BPANN == 'ON':
       print('R_BPANN_avg_100_times:{}'.format(R_BPANN_avg_100_times),file=fp)
       print('RMSE_BPANN_avg_100_times:{}\n'.format(RMSE_BPANN_avg_100_times),file=fp)
    if Model_selection_RF == 'ON':
       print('R_RF_avg_100_times:{}'.format(R_RF_avg_100_times),file=fp)
       print('RMSE_RF_avg_100_times:{}\n'.format(RMSE_RF_avg_100_times),file=fp)
    if Model_selection_KRR == 'ON':
       print('R_KRR_avg_100_times:{}'.format(R_KRR_avg_100_times),file=fp)
       print('RMSE_KRR_avg_100_times:{}\n'.format(RMSE_KRR_avg_100_times),file=fp)
    
   
    print('Testing set:\n', file=fp)
    if Model_selection_GBR == 'ON':
       print('R_GBR_test_set:{}'.format(R_GBR_test_set), file=fp)
       print('RMSE_GBR_test_set:{}\n'.format(RMSE_GBR_test_set),file=fp)
    if Model_selection_SVR == 'ON':
       print('R_SVR_test_set:{}'.format(R_SVR_test_set),file=fp)
       print('RMSE_SVR_test_set:{}\n'.format(RMSE_SVR_test_set),file=fp)
    if Model_selection_BPANN == 'ON':
       print('R_BPANN_test_set:{}\n'.format(R_BPANN_test_set),file=fp)
       print('RMSE_BPANN_test_set:{}'.format(RMSE_BPANN_test_set),file=fp)
    if Model_selection_RF == 'ON':
       print('R_RF_test_set:{}'.format(R_RF_test_set),file=fp)
       print('RMSE_RF_test_set:{}\n'.format(RMSE_RF_test_set),file=fp)
    if Model_selection_KRR == 'ON':
       print('R_KRR_test_set:{}'.format(R_KRR_test_set),file=fp)
       print('RMSE_KRR_test_set:{}\n'.format(RMSE_KRR_test_set),file=fp)

    end = time.time()
    print('run time:{} s'.format(str(end - start)))
    print('run time:{} min'.format(str((end - start) / 60)))
    print('run time:{} s'.format(str(end - start)), file=fp)
    print('run time:{} min'.format(str((end - start) / 60)), file=fp)
    fp.close()

    print('*** Model selection is finished ***')
###########################################################################################################################

#######################################**Hyper_parameter_opt_and_Model_validation**#######################################
if Hyper_parameter_opt_and_Model_validation == 'ON':

    mycopyfile('RhCo_for_model_selection.xlsx', '{}/copy/'.format(path_name))
    os.chdir('{}/copy'.format(path_name))
    os.rename('RhCo_for_model_selection.xlsx', 'RhCo_train_data_for_hyper_para_opt.xlsx')
    os.chdir('{}'.format(path_name))
    mycopyfile('{}/copy/RhCo_train_data_for_hyper_para_opt.xlsx'.format(path_name), '{}/'.format(path_name))
    ############################################### 输入要导入数据文件的名称 ###############################################
    read_excel_name_2 = 'RhCo_train_data_for_hyper_para_opt.xlsx'  # 导入的excel名称
    df_2 = pd.read_excel('{}'.format(read_excel_name_2))
    df_2_head = df_2.columns.to_list()[0:]
    columns_used_in_excel = list(range(len(df_2_head)))  # excel表中有数据的列
    #######################################################################################################################

    mkpath_para_opt = "{}/hyper-parameter-optimization/{}-features-{}-hyper-parameter-opt".format(path_name,columns_used_in_excel[-2],time.strftime("%Y-%m-%d-%H-%M"))
    time_str = time.strftime("%Y-%m-%d-%H-%M")
    mkdir(mkpath_para_opt)
    fp = open('{}/hyper-parameter-optimization/{}-features-{}-hyper-parameter-opt/hyper-parameter-optimization-information.txt'.format(path_name, columns_used_in_excel[-2], time_str), 'a+')
    #######################################################################################################################

    ################################################## 导入数据 #########################################################

    df = pd.read_excel('{}'.format(read_excel_name_2), usecols=columns_used_in_excel)
    y = df['{}'.format(y_column_in_excel)]
    X = df.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)
    #################################################### All parameters opt together ########################################################
    print('***************************************** All parameters opt together result ****************************************************',file=fp)

    print('All hyper parameters optimization ...')

    start = time.time()

    param_grid = {"n_estimators": np.arange(10, 50, 1),
                  "max_depth": np.arange(2, 8, 1),
                  "min_samples_leaf": np.arange(3, 10, 1),
                  "min_samples_split": np.arange(4, 10, 1),
                  "learning_rate": np.arange(0.01, 0.9, 0.01)}
    random.seed(1)
    reg = ensemble.GradientBoostingRegressor()
    GS = EvolutionaryAlgorithmSearchCV(estimator=reg,
                                       params=param_grid,
                                       cv=KFold(n_splits=20),
                                       population_size=50,
                                       gene_mutation_prob=0.01,
                                       gene_crossover_prob=0.03,
                                       generations_number=10,
                                       tournament_size=3,
                                       verbose=1,
                                       n_jobs=1)
    GS.fit(X, y)

    end = time.time()

    print('run time:{} s'.format(str(end - start)), file=fp)
    print('run time:{} min'.format(str((end - start) / 60)), file=fp)

    r, _ = pearsonr(y, GS.predict(X))
    r_estimators_score = r
    Final_best_n_estimators = GS.best_params_.get('n_estimators')
    Final_best_max_depth = GS.best_params_.get('max_depth')
    Final_best_min_samples_leaf = GS.best_params_.get('min_samples_leaf')
    Final_best_min_sample_split = GS.best_params_.get('min_samples_split')
    Final_best_learning_rate = GS.best_params_.get('learning_rate')
    print('r: {}'.format(r), file=fp)
    print('Final_best_n_estimators: {}\n'.format(GS.best_params_.get('n_estimators')), file=fp)
    print('Final_best_max_depth: {}\n'.format(GS.best_params_.get('max_depth')), file=fp)
    print('Final_best_min_samples_leaf: {}\n'.format(GS.best_params_.get('min_samples_leaf')), file=fp)
    print('Final_best_min_sample_split: {}\n'.format(GS.best_params_.get('min_samples_split')), file=fp)
    print('Final_best_learning_rate: {}\n'.format(GS.best_params_.get('learning_rate')), file=fp)
    print('***************************************************************************************************************\n',file=fp)

    ###########################################################################################################################

    ########################################### apply all optimized hyper parameters ########################################

    params = {
        "n_estimators": Final_best_n_estimators,
        "learning_rate": Final_best_learning_rate,
        "max_depth": Final_best_max_depth,
        "min_samples_leaf": Final_best_min_samples_leaf,
        "min_samples_split": Final_best_min_sample_split,
    }

    reg = ensemble.GradientBoostingRegressor(**params)
    reg.fit(X, y)
    y_predict = reg.predict(X)
    r, _ = pearsonr(y, y_predict)
    print("The coefficient of pearsonr on all training set( with all optimized hyper parameters ): {:.4f}".format(r),
          file=fp)
    RMSE = np.sqrt(mean_squared_error(y, reg.predict(X)))
    print("The RMSE on all training set( with all optimized hyper parameters ): {:.4f}".format(RMSE), file=fp)

    save_path = '{}/hyper-parameter-optimization/{}-features-{}-hyper-parameter-opt/'.format(path_name,columns_used_in_excel[-2],time_str)
    joblib.dump(reg, '{}/hyper-parameter-optimization.pkl'.format(save_path))
    X_output = pd.DataFrame(X)
    X_output.to_excel("{}/X_real.xlsx".format(save_path))
    y_output = pd.DataFrame(y)
    y_output.to_excel("{}/y_real.xlsx".format(save_path))
    y_predict_output = pd.DataFrame(y_predict)
    y_predict_output.to_excel("{}/y_predict.xlsx".format(save_path))
    ###########################################################################################################################

    print('######################################################### summry #########################################################',file=fp)
    print('Final_best_n_estimators: {}'.format(Final_best_n_estimators), file=fp)
    print('Final_best_max_depth: {}'.format(Final_best_max_depth), file=fp)
    print('Final_best_min_samples_leaf: {}'.format(Final_best_min_samples_leaf), file=fp)
    print('Final_best_min_sample_split: {}'.format(Final_best_min_sample_split), file=fp)
    print('Final_best_learning_rate: {}'.format(Final_best_learning_rate), file=fp)

    print('*** All hyper parameters optimization are finished ***')
    fp.close()

    # Model_validation
    os.chdir('{}/copy'.format(path_name))
    os.rename( 'RhCo_train_data_for_hyper_para_opt.xlsx','RhCo_train_data_for_model_validation.xlsx')
    mycopyfile('RhCo_train_data_for_model_validation.xlsx', '{}/'.format(path_name))
    os.chdir('{}'.format(path_name))
    train_excel_name = 'RhCo_train_data_for_model_validation.xlsx'  # 导入的excel名称
    df_3 = pd.read_excel('{}'.format(train_excel_name))
    df_3_head = df_3.columns.to_list()[0:]
    sele_head_for_test_data = df_3_head
    columns_used_in_excel = list(range(len(df_3_head)))  # excel表中有数据的列
    read_excel_name_test = "test_data.xlsx"
    df_test = pd.read_excel('{}'.format(read_excel_name_test), usecols=sele_head_for_test_data)
    sele_fea_output = pd.DataFrame(df_test)
    sele_fea_output.to_excel("{}/RhCo_test_data_for_model_validation.xlsx".format(path_name))
    test_data_excel = "{}/RhCo_test_data_for_model_validation.xlsx".format(path_name)
    wb = load_workbook(test_data_excel)
    ws = wb.active
    ws.delete_cols(1)  # 删除第 1列数据
    wb.save(test_data_excel)

    ############################################### 输入要导入数据文件的名称 ###############################################
    test_excel_name = 'RhCo_test_data_for_model_validation.xlsx'  # 导入的excel名称
    #######################################################################################################################

    time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
    mkpath = "{}/Model_validation/{}-features-{}-model-validation".format(path_name, columns_used_in_excel[-2],time_str)
    mkdir(mkpath) 
    #######################################################################################################################

    fp = open('{}/Model_validation/{}-features-{}-model-validation/model-validation-information.txt'.format(path_name,columns_used_in_excel[-2],time_str),'a+')
    print('Read excel name : {}\n'.format(train_excel_name), file=fp)

    print('Model validation ...')

    ################################################## 导入数据 #########################################################

    df = pd.read_excel('{}'.format(train_excel_name), usecols=columns_used_in_excel)
    y_train = df['{}'.format(y_column_in_excel)]
    X_train = df.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)

    df_test = pd.read_excel('{}'.format(test_excel_name), usecols=columns_used_in_excel)
    y_test = df_test['{}'.format(y_column_in_excel)]
    X_test = df_test.drop(['{}'.format(cluster_name_column), '{}'.format(y_column_in_excel)], axis=1)

    ######################################################################################################################

    mkpath = "{}/Model_validation/{}-features-{}-model-validation".format(path_name, columns_used_in_excel[-2],time_str)
    mkdir(mkpath)

    params = {
        "n_estimators": Final_best_n_estimators,
        "learning_rate": Final_best_learning_rate,
        "max_depth": Final_best_max_depth,
        "min_samples_leaf": Final_best_min_samples_leaf,
        "min_samples_split": Final_best_min_sample_split,
    }

    reg = ensemble.GradientBoostingRegressor(**params)
    reg.fit(X_train, y_train)
    r_train, _ = pearsonr(y_train, reg.predict(X_train))
    RMSE_train = np.sqrt(mean_squared_error(y_train, reg.predict(X_train)))

    r, _ = pearsonr(y_test, reg.predict(X_test))
    RMSE = np.sqrt(mean_squared_error(y_test, reg.predict(X_test)))

    save_path = '{}/Model_validation/{}-features-{}-model-validation'.format(path_name, columns_used_in_excel[-2],time_str)

    joblib.dump(reg, '{}/Model-validation-output.pkl'.format(save_path))
    X_train_output = pd.DataFrame(X_train)
    X_train_output.to_excel("{}/X_train_real.xlsx".format(save_path))
    y_train_output = pd.DataFrame(y_train)
    y_train_output.to_excel("{}/y_train_real.xlsx".format(save_path))
    y_train_predict_output = pd.DataFrame(reg.predict(X_train))
    y_train_predict_output.to_excel("{}/y_train_predict.xlsx".format(save_path))
    X_test_output = pd.DataFrame(X_test)
    X_test_output.to_excel("{}/X_test_real.xlsx".format(save_path))
    y_test_output = pd.DataFrame(y_test)
    y_test_output.to_excel("{}/y_test_real.xlsx".format(save_path))
    y_test_predict_output = pd.DataFrame(reg.predict(X_test))
    y_test_predict_output.to_excel("{}/y_test_predict.xlsx".format(save_path))

    ############################################ 绘制散点图 ##########################################################
    ###*绘制训练集上预测值和真实值之间的散点图
    size = len(y_train)
    Y = np.linspace(-8.5,-14.5,size)
    X = np.linspace(-8.5,-14.5,size)
    fig = plt.figure(dpi=300,figsize=(20,20))
    ax1 = fig.add_subplot(1,1,1)
    ax1.set_title("Train set", size=45)
    ax1.scatter(y_train,reg.predict(X_train), marker='*', c='red',s=1000,alpha=0.7)
    plt.xticks(fontproperties='Times New Roman', size=45)
    plt.yticks(fontproperties='Times New Roman', size=45)
    ax1.set_xlabel("Experimental lgk1", size=45)
    ax1.set_ylabel("Predicted lgk1", size=45)
    ax1.plot(X,Y)
    plt.savefig('{}/Train_set.png'.format(save_path))  

    ###*绘制测试集上预测值和真实值之间的散点图
    size = len(y_test)
    Y = np.linspace(-8.5,-14.5,size)
    X = np.linspace(-8.5,-14.5,size)
    fig = plt.figure(dpi=300,figsize=(20,20))
    ax1 = fig.add_subplot(1,1,1)
    ax1.set_title("Test set", size=45)
    ax1.scatter(y_test,reg.predict(X_test), marker='*', c='red',s=1000,alpha=0.7)
    plt.xticks(fontproperties='Times New Roman', size=45)
    plt.yticks(fontproperties='Times New Roman', size=45)
    ax1.set_xlabel("Experimental lgk1", size=45)
    ax1.set_ylabel("Predicted lgk1", size=45)
    ax1.plot(X,Y)
    plt.savefig('{}/Test_set.png'.format(save_path)) 

    ########################################################################################################################################

    print('**************************************************** summry *****************************************************',file=fp)
    print('n_estimators:{}'.format(Final_best_n_estimators), file=fp)
    print('max_depth:{}'.format(Final_best_max_depth), file=fp)
    print('min_samples_leaf:{}'.format(Final_best_min_samples_leaf), file=fp)
    print('best_min_sample_split:{}'.format(Final_best_min_sample_split), file=fp)
    print('learning_rate:{}\n'.format(Final_best_learning_rate), file=fp)

    print("The coefficient of pearsonr on train set: {:.4f}".format(r_train), file=fp)
    print("The RMSE on train set: {:.4f}\n".format(RMSE_train), file=fp)
    print("* The coefficient of pearsonr on test set: {:.4f}".format(r), file=fp)
    print("* The RMSE on test set: {:.4f}".format(RMSE), file=fp)

    if Feature_selection_LOOCV  == 'ON':
       dst_dir_info_featu_sele = '{}'.format(mkpath_fea_sele_loocv)
       shutil.move('RhCo_train_data_for_feature_slection.xlsx',dst_dir_info_featu_sele)

    if Feature_selection_stat   == 'ON':
       dst_dir_info_featu_sele = '{}'.format(mkpath_featu_sele_stat)
       shutil.move('RhCo_train_data_for_feature_slection.xlsx',dst_dir_info_featu_sele)
       
    if Model_selection    == 'ON':
       dst_dir_info_mode_sele = '{}'.format(mkpath_model_sele)
       shutil.move('RhCo_for_model_selection.xlsx',dst_dir_info_mode_sele)

    if Hyper_parameter_opt_and_Model_validation == 'ON':    
       dst_dir_info_para_opt = '{}'.format(mkpath_para_opt)
       shutil.move('RhCo_train_data_for_hyper_para_opt.xlsx',dst_dir_info_para_opt)

       dst_dir_info = '{}/Model_validation/{}-features-{}-model-validation'.format(path_name, columns_used_in_excel[-2],time_str)
       shutil.move('RhCo_train_data_for_model_validation.xlsx',dst_dir_info)     
       shutil.move('RhCo_test_data_for_model_validation.xlsx',dst_dir_info)
    
    print('***Model validation are finished***')
    fp.close()
    copy_path = '{}/copy'.format(path_name)
    shutil.rmtree(copy_path,ignore_errors=True)


