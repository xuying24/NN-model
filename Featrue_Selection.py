########################### 代码功能：利用GBR特征重要性和末尾淘汰机制进行特征选择####################################
################################################## 导入模块 #########################################################

import pandas as pd
import matplotlib.pyplot as plt
from pandas import Series
import numpy as np
from sklearn import ensemble,preprocessing
from sklearn.preprocessing import minmax_scale,StandardScaler, Normalizer
from sklearn.inspection import permutation_importance
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split,LeaveOneOut
from sklearn.metrics import r2_score
from scipy.stats import pearsonr
from sklearn.feature_selection import SelectFromModel
import seaborn as sb
import os
import sys
import time
from openpyxl import load_workbook

############################################### 输入要导入数据文件的名称 ###############################################
path_name       = os.getcwd()
file_name       = 'RhCo--featrue_sele_r_method-loocv.py'
read_excel_name = 'train_data_for_feature_selection.xlsx'                          # 导入的excel名称
columns_used_in_excel = [i for i in range(0,14)]        # excel表中有数据的列
y_column_in_excel = 'lgk1'                                            # excel表中目标值那一列的title
cluster_name_column = 'Cluster'                                        # excel表中团簇名称那一列的title
#######################################################################################################################

########################################### 定义新建文件夹函数 ########################################################
def mkdir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)
    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)
        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False
# 定义要创建的目录
mkpath = "{}/feature_selection/loocv-{}-feature-selection".format(path_name, time.strftime("%Y-%m-%d-%H-%M-%S"))
time_str = time.strftime("%Y-%m-%d-%H-%M-%S")
mkdir(mkpath)
#######################################################################################################################


fp=open( '{}/feature_selection/loocv-{}-feature-selection/feature-select-information.txt'.format(path_name, time_str), 'a+')


############################################ Define copy function ######################################################
# srcfile 需要复制、移动的文件   
# dstpath 目的地址
import glob
import shutil

def mycopyfile(srcfile,dstpath):                       # 复制函数
    if not os.path.isfile(srcfile):
        print ("%s not exist!"%(srcfile))
    else:
        fpath,fname=os.path.split(srcfile)             # 分离文件名和路径
        if not os.path.exists(dstpath):
            os.makedirs(dstpath)                       # 创建路径
        shutil.copy(srcfile, dstpath + fname)          # 复制文件
        print ("copy %s -> %s"%(srcfile, dstpath + fname))

src_dir_info = '{}/'.format(path_name)

dst_dir_info = '{}/feature_selection/loocv-{}-feature-selection/'.format(path_name, time_str)

src_file_list =  glob.glob(src_dir_info + '{}'.format(read_excel_name))

src_file_list2 =  glob.glob(src_dir_info + '{}'.format(file_name))

for srcfile in src_file_list:
    mycopyfile(srcfile, dst_dir_info)
    
for srcfile in src_file_list2:
    mycopyfile(srcfile, dst_dir_info)
#######################################################################################################################
  

    
################################################## 导入数据 #########################################################

df                  = pd.read_excel('{}'.format(read_excel_name),usecols=columns_used_in_excel)
print(df.describe())
print('Read excel name : {}\n'.format(read_excel_name),file=fp)

X_name          =   df.columns.to_list()[1:-1]

print('initial features in database:\n',file=fp)
print(X_name,'\n',file=fp)

y                   = df['{}'.format(y_column_in_excel)]
X                   = df.drop([cluster_name_column] + [y_column_in_excel], axis=1)


initial_num_feature = len(X_name)

######################################################################################################################
  
############################################## 定义存放数据的数组 #####################################################

#************************************* 存放【特征选择】过程中的重要数据 ***********************************************
R2_scorel_feature        = []                                     # 用于存放特征选择每次循环中的R2数值
r_scorel_feature         = []                                     # 用于存放特征选择每次循环中的r数值
RMSE_scorel_feature      = []                                     # 用于存放特征选择每次循环中的RMSE数值
num_feature              = []                                     # 用于存放特征选择每次循环中的当前特征数目
NUN_array = []
#**********************************************************************************************************************

   
################################################### 特征选择 ##########################################################
print('#################################################### Feature selection ##################################################',file=fp)
#***************************************** 利用循环语句删除不重要特征 #################################################
Min_num_featrue = 1
num_cycle = (initial_num_feature - Min_num_featrue)  + 1                      # 设置循环次数
r_scorel_arry = [NUN_array] * (initial_num_feature)
features_arry = [NUN_array] * (initial_num_feature)

for i in range(0,num_cycle):
    X_name_sorted_idx = []

    X_dict = X.values
    y_dict = y.values
    
    print('####################################################{} features to selected ##################################################'.format(initial_num_feature - i),file=fp)
    X_name_sorted_idx        = []                                             # 定义存放根据特征重要性排好序的特征名的数组
    y_test_real_all = []
    y_test_predict_all = []
    
    loo = LeaveOneOut()
    loo.get_n_splits(X_dict)
    for train_index, test_index in loo.split(X_dict):
        X_train_loo, X_test_loo = X_dict[train_index], X_dict[test_index]
        y_train_loo, y_test_loo = y_dict[train_index], y_dict[test_index]
        y_test_real_all.append(y_test_loo)
        
        reg = ensemble.GradientBoostingRegressor()                                # 定义模型名称
        reg.fit(X_train_loo, y_train_loo)                                                 # 模型拟合
        y_test_predict = reg.predict(X_test_loo)
        y_test_predict_all.append(y_test_predict)
        
    y_test_real_all = np.squeeze(y_test_real_all)
    y_test_predict_all = np.squeeze(y_test_predict_all)
        
    R2 = r2_score(y_test_real_all, y_test_predict_all)                                # 在测试集上计算决定系数
    r,_ = pearsonr(y_test_real_all, y_test_predict_all)                               # 在测试集上计算皮尔逊相关系数
    r_scorel_arry[i] = r
    
    RMSE = np.sqrt(mean_squared_error(y_test_real_all, y_test_predict_all))          # 在测试集上计算均方根误差

    print("The coefficient of R2 (including {} features) : {:.4f}".format((initial_num_feature - i), R2),file=fp)
    print("The coefficient of pearsonr (including {} features): {:.4f}".format((initial_num_feature - i), r),file=fp)
    print("The RMSE(including {} features): {:.4f}".format((initial_num_feature - i), RMSE),file=fp)
    
    feature_importance = reg.feature_importances_                             # 利用GBR方法计算得到各特征的重要性
    print(feature_importance, file=fp)
    sorted_idx = np.argsort(feature_importance)                               # 利用GBR方法计算得到各特征重要性的排序索引
    feature_importance_sorted = feature_importance[sorted_idx]
    
    for j in sorted_idx:
        X_name_sorted_idx.append(X_name[j])
    features_arry[i] = X_name_sorted_idx

    
    num_feature.append(X.shape[1])
    R2_scorel_feature.append(R2)
    r_scorel_feature.append(r)
    ############################################## 绘制特征重要性的条形图 #########################################

    pos = np.arange(sorted_idx.shape[0]) + 0.5
    fig = plt.figure(figsize=(12, 12))
    plt.subplot(1, 1, 1)
    plt.barh(pos, feature_importance_sorted, height=0.8, align="center")
    plt.xticks(fontproperties = 'Times New Roman', size = 14)
    plt.yticks(pos, np.array(X_name)[sorted_idx], fontproperties = 'Times New Roman', size = 14)
    plt.title("Feature Importance (MDI): r = {:.4f}, R2 = {:.4f}, RMSE = {:.4f}".format(r,R2,RMSE), fontproperties = 'Times New Roman', size = 16)
    plt.savefig('{}/feature_selection/loocv-{}-feature-selection/Feature_Importance_{}_features.png'.format(path_name,time_str,(initial_num_feature - i)))
    
    ######################################### 绘制所有特征之间相关性的热力图 #########################################
    
    fig = plt.figure(figsize=(12, 12))
    plt.subplot(1, 1, 1)
    plt.xticks(fontproperties = 'Times New Roman', size = 14)
    plt.yticks(fontproperties = 'Times New Roman', size = 14)
    #dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlGnBu", annot=True)
    dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlGnBu", annot=True)#蓝绿色
    #dataplot = sb.heatmap(X.corr(method ='pearson'), cmap="YlOrRd")#红橙色
    plt.savefig('{}/feature_selection/loocv-{}-feature-selection/{}_features_heatmap.png'.format(path_name, time_str,(initial_num_feature - i)))        
    #plt.show()

    #######################################################################################################################
    
    X.drop(columns=X.columns[[sorted_idx[0]]],inplace = True)                        # 特征按照重要性进行排序后，将重要性最低的特征从数据列表中删除
    del X_name[sorted_idx[0]]                                                # 将对应的特征名称一并删除
                                                                    
#**********************************************************************************************************************

#######################################################################################################################
    
################################### 绘制决定系数和皮尔逊相关系数与特征数目的关系图 ####################################
fig = plt.figure(figsize=(12, 4))
plt.subplot(1,1, 1)
plt.xticks(fontproperties = 'Times New Roman', size = 14)
plt.yticks(fontproperties = 'Times New Roman', size = 14)
plt.xlabel('Feature number',fontproperties = 'Times New Roman', size = 14)
plt.ylabel('R2 of GBR algorithm',fontproperties = 'Times New Roman', size = 14)
plt.plot(num_feature,R2_scorel_feature)
plt.savefig('{}/feature_selection/loocv-{}-feature-selection/R2_vs_Feature_number_{}.png'.format(path_name,time_str, time_str))  
#plt.show()

plt.figure(figsize=[12,4])
plt.subplot(1, 1, 1)
plt.xticks(fontproperties = 'Times New Roman', size = 14)
plt.yticks(fontproperties = 'Times New Roman', size = 14)
plt.xlabel('Feature number',fontproperties = 'Times New Roman', size = 14)
plt.ylabel('r of GBR algorithm',fontproperties = 'Times New Roman', size = 14)
plt.plot(num_feature,r_scorel_feature)
plt.savefig('{}/feature_selection/loocv-{}-feature-selection/r_vs_Feature_number_{}.png'.format(path_name,time_str, time_str))  
#plt.show()

# 读取非统计法选择的特征
selected_featu = features_arry[r_scorel_arry.index(max(r_scorel_arry))]
sele_head = [cluster_name_column] + [y_column_in_excel]  + selected_featu 

# 选择Excel表中的重要特征列，然后生成RhCo_for_model_selection.xisx
os.chdir('{}'.format(path_name))
df = pd.read_excel('{}'.format(read_excel_name), usecols=columns_used_in_excel, sheet_name='Sheet1')
all_head_name = df.columns.to_list()[0:]
df_1 = pd.read_excel('{}'.format(read_excel_name), usecols=sele_head)
sele_fea_output = pd.DataFrame(df_1)
sele_fea_output.to_excel("{}/Train_data_for_model_validation.xlsx".format(path_name))
model_validation_train_data_excel = "{}/Train_data_for_model_validation.xlsx".format(path_name)
wb = load_workbook(model_validation_train_data_excel)
ws = wb.active
ws.delete_cols(1)  # 删除第 1列数据
wb.save(model_validation_train_data_excel)

df_test_data = pd.read_excel('test_data.xlsx', usecols=sele_head)
test_data_sele_fea_output = pd.DataFrame(df_test_data)
test_data_sele_fea_output.to_excel("{}/Test_data_for_model_validation.xlsx".format(path_name))
model_validation_test_data_excel = "{}/Test_data_for_model_validation.xlsx".format(path_name)
wb_test_data = load_workbook(model_validation_test_data_excel)
ws_test_data = wb_test_data.active
ws_test_data.delete_cols(1)  # 删除第 1列数据
wb_test_data.save(model_validation_test_data_excel)
########################################################################################################################
print('*** Feature selection finished ***')
fp.close()